import numpy as np
import math
from scipy import signal

from openpyxl import Workbook
from openpyxl import load_workbook
#from openpyxl.compat import range
import openpyxl.compat

import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2TkAgg
from matplotlib.backend_bases import key_press_handler
from matplotlib.figure import Figure

import tkinter as tk
from tkinter import ttk

from tkinter.filedialog import askopenfilename, asksaveasfilename

import os
from datetime import datetime


class Filterer(object):

    def __init__(self):
        pass


    def gaussianFilter(self, x, window_size=10, std=7):
        kernel = signal.gaussian(window_size, std=std)

        x_origin = np.copy(x)
        x_result = np.zeros(x.shape)

        for i, value in enumerate(x_origin):
            offset = math.floor(window_size/2.0)
            
            first_idx = i-offset
            if first_idx < 0:
                first_idx = 0


            src = x_origin[first_idx : i+offset +1] 
         
            if len(src) != len(kernel):
                x_result[i] = x_origin[i]
            elif len(src) == len(kernel):
                x_result[i] = np.sum( src * kernel / float(window_size))

        return x_result


    def averageFilter(self, x, window_size=3):

        x_origin = np.copy(x)
        x_result = np.zeros(x.shape)

        for i, value in enumerate(x_origin):
            offset = math.floor(window_size/2.0)
            
            first_idx = i-offset
            if first_idx < 0:
                first_idx = 0

            src = x_origin[first_idx: i+offset +1]
            if len(src) != window_size:
                x_result[i] = x_origin[i]
            else:
                x_result[i] = np.sum( src / float(window_size))

        return x_result

    def findPeak(self, x ):
        
        x_result = np.zeros(x.shape)

        for i in range(1, len(x)-1):
            if x[i] > x[i-1] and x[i] >x[i+1]:
                x_result[i] = 1
        
        return x_result

class XlHandler(object):

    def __init__(self):
        self.wb = None
    

    def getDataFrom(self, start, end):
        return 

    def loadFile(self):
        pass
    def saveFile(self):
        pass


class GraphTool(object):

    def __init__(self):
        self.mode = "gaussian"
        self.wb = None
        self.initGui()

    def initGui(self):
        self.text_size = 6
        self.sheet_max_num_in_row = 8

        self.root = tk.Tk()
        self.root.wm_title("Graph Tool Controller")
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)
        self.initControlFrame(self.root)
        self.initGraphFrame(self.root)

        
        for child in self.controlframe.winfo_children():
            child.grid_configure(sticky=(tk.W, tk.E ))
        #for child in self.graphframe.winfo_children():
        #    child.grid_configure(sticky=(tk.W, tk.E, tk.N, tk.S ))

        self.root.bind('<Return>', lambda event, i=self: i.process())

    def initControlFrame(self, root):
        ###
        # controlframe
        controlframe = ttk.Frame(root)
        
        controlroot = ttk.Frame(root)
        controlroot.grid(row=0, column=0, sticky=(tk.N, tk.W, tk.E))

        menuframe = ttk.Frame(controlroot)
        controlframe = ttk.Frame(controlroot)
        optionframe = ttk.Frame(controlroot)
        sheetframe = ttk.Frame(controlroot)        
        
        #controlframe.columnconfigure(0, weight=1)
        #controlframe.rowconfigure(0, weight=1)
        menuframe.grid(row=0,column=0, sticky=(tk.W))
        controlframe.grid(row=2, column=0, sticky=(tk.N, tk.W))
        optionframe.grid(row=2,column=2, sticky=(tk.E))
        sheetframe.grid(row=1, column=0, columnspan=self.sheet_max_num_in_row+1, sticky=(tk.W,tk.E))
        
        # controlframe column configure
        #for i in openpyxl.compat.range(4):
        #    controlframe.columnconfigure(i, weight=3%(i+1) )

        ### menuframe
        ttk.Button(menuframe, text="open file", command=self.openFile).grid(row=0, column=0)
        ttk.Button(menuframe, text="save file", command=self.saveFile).grid(row=0, column=1)

        self.menuframe = menuframe
        
        ### controlframe

        self.x_start_var = tk.StringVar()
        self.x_end_var = tk.StringVar()

        self.y_start_var = tk.StringVar()
        self.y_end_var = tk.StringVar()

        # x variable
        ttk.Label(controlframe, text="x start").grid(row=0, column=0)
        ttk.Label(controlframe, text="x end").grid(row=0, column=2)
        x_start_entry = ttk.Entry(controlframe, textvariable=self.x_start_var, width=self.text_size)
        x_start_entry.grid(row=0, column=1)
        x_start_entry.focus()
        ttk.Entry(controlframe, textvariable=self.x_end_var, width=self.text_size).grid(row=0, column=3)

        # y variable
        ttk.Label(controlframe, text="y start").grid(row=1, column=0)
        ttk.Label(controlframe, text="y end").grid(row=1, column=2)
        ttk.Entry(controlframe, textvariable=self.y_start_var, width=self.text_size).grid(row=1, column=1)
        ttk.Entry(controlframe, textvariable=self.y_end_var, width=self.text_size).grid(row=1, column=3)

        # Run button

        self.controlframe = controlframe

    
        self.current_sheet_text = tk.StringVar()
        self.current_sheet_label = tk.Label(sheetframe, textvariable=self.current_sheet_text)
        self.current_sheet_label.grid(row=0, column=0, sticky=(tk.W,tk.E))

        self.current_sheet_text.set("sheet name")
        self.sheetframe = sheetframe

        ## 
        # option Frame
    
        ## real_time_frame
        real_time_frame=ttk.Frame(optionframe)
        self.real_time_flag = tk.IntVar()
        ttk.Checkbutton(real_time_frame, text="real time", variable=self.real_time_flag).grid(row=0,column=0)
        ttk.Button(real_time_frame, text="run", command=self.process).grid(row=1, column=0) 
        ttk.Label(real_time_frame, text="click run or enter").grid(row=2, column=0)
        self.real_time_flag.set(0)

        ## graph_limit_frame
        graph_limit_frame = ttk.Frame(optionframe)
        self.graph_limit_flag = tk.IntVar()
        self.graph_max_y = tk.DoubleVar()
        self.graph_min_y = tk.DoubleVar()
        ttk.Checkbutton(graph_limit_frame, text="graph limit", variable=self.graph_limit_flag).grid(row=0, column=0)
        ttk.Label(graph_limit_frame, text="max y").grid(row=1, column=0)
        ttk.Entry(graph_limit_frame, textvariable=self.graph_max_y, width=self.text_size).grid(row=1, column=1)
        ttk.Label(graph_limit_frame, text="min y").grid(row=2, column=0)
        ttk.Entry(graph_limit_frame, textvariable=self.graph_min_y, width=self.text_size).grid(row=2, column=1)
        self.graph_limit_flag.set(1)
        self.graph_max_y.set(140)
        self.graph_min_y.set(0)

        ## filter_original_frame
        filter_original_frame = ttk.Frame(optionframe)
        #filter_original_frame.grid(row=0, column=0)
        self.original_flag = tk.IntVar()
        ttk.Checkbutton(filter_original_frame, text="original", variable=self.original_flag).grid(row=0, column=0)
        self.original_flag.set(1)
        
        for child in filter_original_frame.winfo_children():
            child.grid_configure(sticky=(tk.W, tk.N))

        ## filter_gaussian_frame
        filter_gaussian_frame = ttk.Frame(optionframe)
        #filter_gaussian_frame.grid(row=0, column=1)
        self.gaussian_flag = tk.IntVar()
        self.gaussian_std = tk.DoubleVar()
        self.gaussian_window_size = tk.IntVar()
        ttk.Checkbutton(filter_gaussian_frame, text="gaussian filter", variable=self.gaussian_flag).grid(row=0,column=0)
        ttk.Label(filter_gaussian_frame, text="window size").grid(row=1, column=0)
        ttk.Entry(filter_gaussian_frame, textvariable=self.gaussian_window_size, width=self.text_size).grid(row=1, column=1)
        ttk.Label(filter_gaussian_frame, text="std").grid(row=2, column=0)
        ttk.Entry(filter_gaussian_frame, textvariable=self.gaussian_std, width=self.text_size).grid(row=2, column=1)

        self.gaussian_flag.set(0)
        self.gaussian_std.set(3)
        self.gaussian_window_size.set(3)

        for child in filter_gaussian_frame.winfo_children():
            child.grid_configure(sticky=(tk.W, tk.N))

        ## filter_average_frame
        filter_average_frame = ttk.Frame(optionframe)
        #filter_average_frame.grid(row=0, column=2)
        self.average_flag = tk.IntVar()
        self.average_window_size = tk.IntVar()
        ttk.Checkbutton(filter_average_frame, text="average filter", variable=self.average_flag).grid(row=0,column=0)
        ttk.Label(filter_average_frame, text="window size").grid(row=1, column=0)
        ttk.Entry(filter_average_frame, textvariable=self.average_window_size, width=self.text_size).grid(row=1, column=1) 

        self.average_flag.set(0)
        self.average_window_size.set(3)

        for child in filter_average_frame.winfo_children():
            child.grid_configure(sticky=(tk.W, tk.N))


        for i, child in enumerate(optionframe.winfo_children()):
            child.grid_configure(row=0, column=i, sticky=(tk.W, tk.N))


    def initGraphFrame(self, root):
        ###
        # graphframe
        graphframe = ttk.Frame(root)
        graphframe.grid(row=1, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        self.figure = Figure()

        canvas = FigureCanvasTkAgg(self.figure, master=graphframe)
        canvas.show()
        canvas.get_tk_widget().grid(row=0, column=0)

        toolbar = NavigationToolbar2TkAgg(canvas, graphframe)
        toolbar.update()
        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)

        self.canvas = canvas
        self.toolbar = toolbar

        self.graphframe = graphframe

    def update(self):
        if self.real_time_flag.get() == 1:
            self.process()

        self.root.after(100, self.update)

    def draw(self):
        pass

    def process(self):
        try:
            filterer = Filterer()

            x_data = self.ws[self.x_start_var.get():self.x_end_var.get()]
            y_data = self.ws[self.y_start_var.get():self.y_end_var.get()]

            x = []
            y = []

            for row in x_data:
                for cell in row:
                    x.append(cell.value) 
            for row in y_data:
                for cell in row:
                    y.append(cell.value)

            # data read part
            x = np.array(x)
            y = np.array(y)



            self.figure.clear()
            f = self.figure.add_subplot(111)

            if self.original_flag.get() == 1:
                f.plot(x,y,color='black', label='original')
            if self.average_flag.get() == 1:
                y_filtered_with_average     = filterer.averageFilter(y,window_size=self.average_window_size.get())
                f.plot(x,y_filtered_with_average, color="green",label='average filter')
            if self.gaussian_flag.get() == 1:
                y_filtered_with_gaussian    = filterer.gaussianFilter(y, window_size=self.gaussian_window_size.get(), std=self.gaussian_std.get())
                f.plot(x,y_filtered_with_gaussian, color="red",label='gaussian filter')

            if self.graph_limit_flag.get() == 1:
                f.set_ylim([self.graph_min_y.get(), self.graph_max_y.get()])

            # legend
            f.legend(loc='upper left', frameon=False)
       
            self.canvas.show()
            #self.toolbar.update()
        except:
            pass
        

    def openFile(self):
        file_path = askopenfilename(#initialdir="~/",
                               filetypes =(("Excel Files", "*.xlsx"),("All Files","*.*")),
                               title = "Choose a file."
                               )
        # when cancel the file dialog 
        if(file_path == ''):
            return

        self.wb = load_workbook(file_path, data_only=True)

        for i, child in enumerate(self.sheetframe.winfo_children()):
            if i != 0:
                child.destroy()
        self.makeSheetBtn()

    def saveFile(self):
        file_path = asksaveasfilename( defaultextension=".xlsx")         
        # TODO : add logger
        if file_path == None:
            return
       
        wb = None
        ws = None
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            result_title = "result_"+str(datetime.now().year)+"_"+str(datetime.now().month)+"_"+str(datetime.now().day)+"_"+str(datetime.now().hour)+"_"+str(datetime.now().minute)+"_"+str(datetime.now().second)
            ws = wb.create_sheet(title=result_title)
        else:
            wb = Workbook()
            ws =wb.active
        self.fillResult(ws)

        wb.save(file_path)

    def fillResult(self, ws):
        filterer = Filterer()

        x_data = self.ws[self.x_start_var.get():self.x_end_var.get()]
        y_data = self.ws[self.y_start_var.get():self.y_end_var.get()]

        x = []
        y = []

        for row in x_data:
            for cell in row:
                x.append(cell.value) 
        for row in y_data:
            for cell in row:
                y.append(cell.value)

        # data read part
        x = np.array(x)
        y = np.array(y)
        

        ## memory allocation
        for row in openpyxl.compat.range(1, len(y)+3):
            for col in openpyxl.compat.range(1,5):
                ws.cell(row=row,column=col) 

        offset = 2

        ## save x column
        col_name_x = 'A'
        start_x = col_name_x+str(offset)
        end_x = col_name_x+str(len(x)+offset-1)
        A = ws[start_x:end_x]
        ws['A1']='x'

        for i, row in enumerate(A):
            for cell in row:
                cell.value = x[i]


        def fillCol(ws, col_name, field_name,offset, data):

            ## save original
            col_name = col_name
            start_y = col_name+str(offset)
            end_y = col_name+str(len(data)+offset-1)
            col = ws[start_y:end_y]
            ws[col_name+'1']= field_name
            for i, row in enumerate(col):
                for cell in row:
                    cell.value = data[i]



        ## save original
        col_name_y = 'B'
        start_y = col_name_y+str(offset)
        end_y = col_name_y+str(len(x)+offset-1)
        B = ws[start_y:end_y]
        ws['B1']='y origin'

        for i, row in enumerate(B):
            for cell in row:
                cell.value = y[i]

        ## save peak of original
        peak_y = filterer.findPeak(y)
        print(peak_y)
        fillCol(ws, 'C', 'peak y origin', offset, peak_y) 

        ## gaussian 
        y_filtered_with_gaussian    = filterer.gaussianFilter(y, window_size=self.gaussian_window_size.get(), std=self.gaussian_std.get())

        ## save original
        col_name_y_gaussian = 'D'
        start_y_gaussian = col_name_y_gaussian+str(offset)
        end_y_gaussian = col_name_y_gaussian+str(len(x)+offset-1)
        D = ws[start_y_gaussian:end_y_gaussian]
        ws['D1']='y filtered with gaussian kernel'

        for i, row in enumerate(D):
            for cell in row:
                cell.value = y_filtered_with_gaussian[i]

        ## save peak of gaussian 
        peak_y_gaussian = filterer.findPeak(y_filtered_with_gaussian)
        fillCol(ws, 'E', 'peak y gaussian', offset, peak_y_gaussian) 


        y_filtered_with_average     = filterer.averageFilter(y,window_size=self.average_window_size.get())

        ## save original
        col_name_y_average = 'F'
        start_y_average = col_name_y_average+str(offset)
        end_y_average = col_name_y_average+str(len(x)+offset-1)
        F = ws[start_y_average:end_y_average]
        ws['F1']='y filtered with average kernel'

        for i, row in enumerate(F):
            for cell in row:
                cell.value = y_filtered_with_average[i]
      
        ## save peak of average
        peak_y_average = filterer.findPeak(y_filtered_with_average)
        fillCol(ws, 'G', 'peak y average', offset, peak_y_average) 


    def makeSheetBtn(self):
        sheet_names = self.wb.get_sheet_names()


        for i, sheet_name in enumerate(sheet_names):
            tmp = sheet_name
            ttk.Button(self.sheetframe, text=sheet_name, command=lambda sheet_name=sheet_name: self.selectSheet(sheet_name)).grid(row=math.floor(i/self.sheet_max_num_in_row), column=(i+1)%self.sheet_max_num_in_row,sticky=(tk.W)) 

    def selectSheet(self, sheet_name):
        self.current_sheet_text.set(sheet_name)
        self.ws = self.wb[sheet_name]

    def run(self):
        self.root.after(100, self.update)
        self.root.mainloop()

if __name__=='__main__':
    GraphTool().run()
    #main()
    #onlyOneFilter()
